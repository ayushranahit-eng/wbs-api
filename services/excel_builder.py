import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from core.config import settings

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
HEADER_FILL = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
PHASE_FILL = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")
TITLE_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _auto_width(ws, min_col, max_col):
    for col in range(min_col, max_col + 1):
        letter = get_column_letter(col)
        max_len = max(
            (len(str(c.value)) for c in ws[letter] if c.value), default=0
        )
        ws.column_dimensions[letter].width = max(8, min(max_len + 3, 45))


def _apply_styles(ws, max_col):
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = CENTER

    for row in range(6, ws.max_row + 1):
        a = ws[f"A{row}"]
        b = ws[f"B{row}"]
        if a.value:
            a.font = Font(bold=True)
        if b.value and b.__class__.__name__ != "MergedCell":
            b.font = Font(bold=True)
        if a.value and str(a.value).startswith("PHASE"):
            a.font = Font(bold=True, size=14)
            a.fill = PHASE_FILL
            a.alignment = LEFT


def _insert_phase_headers(ws, max_col):
    current_phase = None
    phase_counter = 1
    row_num = 6

    while row_num <= ws.max_row:
        val = ws[f"A{row_num}"].value
        if not val:
            row_num += 1
            continue
        if val != current_phase:
            ws.insert_rows(row_num)
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=max_col)
            cell = ws.cell(row=row_num, column=1)
            cell.value = f"PHASE {phase_counter} : {val}"
            cell.font = Font(bold=True, size=14)
            cell.fill = PHASE_FILL
            ws.row_dimensions[row_num].height = 26
            current_phase = val
            phase_counter += 1
            row_num += 1
        row_num += 1


def _dedup_column(ws, col_letter, skip_keyword="PHASE"):
    prev = None
    for row in range(6, ws.max_row + 1):
        cell = ws[f"{col_letter}{row}"]
        if cell.__class__.__name__ == "MergedCell":
            continue
        if cell.value and skip_keyword in str(cell.value):
            prev = None
            continue
        if cell.value == prev:
            cell.value = ""
        elif cell.value:
            prev = cell.value


def _write_header(ws, project_title, company_name, project_manager, today_date, max_col):
    ws.merge_cells(f"A1:{get_column_letter(max_col)}1")
    ws["A1"] = "PROJECT TIMELINE (W.B.S)"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = LEFT

    for cell in ws[1]:
        cell.fill = TITLE_FILL

    ws["A2"] = "PROJECT TITLE";  ws["B2"] = project_title
    ws["D2"] = "COMPANY NAME";   ws["E2"] = company_name
    ws["A3"] = "PROJECT MANAGER"; ws["B3"] = project_manager
    ws["D3"] = "DATE";            ws["E3"] = today_date

    bold11 = Font(bold=True, size=11)
    for c in ["B2", "E2", "B3", "E3"]:
        ws[c].font = bold11

    for cell in ws[5]:
        cell.font = Font(bold=True, size=11)
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def build_excel(
    all_rows: list,
    project_title: str,
    company_name: str,
    project_manager: str,
    job_id: str,
) -> tuple[str, str]:
    """Returns (full_wbs_path, sales_wbs_path)"""

    os.makedirs(settings.OUTPUT_FOLDER, exist_ok=True)
    today = datetime.now().strftime("%d-%b-%Y")

    full_cols = [
        "Phase", "Module Name", "Task ID", "Task Title", "Sub Task ID", "Sub Task",
        "Priority", "Dependency", "Owner", "Status", "Start Date",
        "Working Duration (Days)", "Actual Completion Date", "Delay (In Days)",
        "% Complete", "Effort (Hours)", "Sprint / Milestone", "Remarks / Comments",
    ]
    sales_cols = ["Phase", "Module Name", "Task ID", "Task Title", "Sub Task ID", "Sub Task", "Remarks / Comments"]

    df = pd.DataFrame(all_rows)[full_cols].fillna("")

    def _safe(name):
        return "".join(c if c.isalnum() or c in " _-" else "" for c in name).strip().replace(" ", "_")

    date_str = datetime.now().strftime("%d%b%Y_%H%M")
    base = f"{_safe(company_name)}_{_safe(project_title)}_{date_str}"

    full_path = os.path.join(settings.OUTPUT_FOLDER, f"WBS_{base}.xlsx")
    sales_path = os.path.join(settings.OUTPUT_FOLDER, f"Sales_WBS_{base}.xlsx")

    # --- Full WBS ---
    df.to_excel(full_path, sheet_name="WBS", index=False, startrow=4)
    wb = load_workbook(full_path)
    ws = wb["WBS"]
    _write_header(ws, project_title, company_name, project_manager, today, 18)
    _insert_phase_headers(ws, 18)
    _dedup_column(ws, "A")
    _dedup_column(ws, "B")
    _apply_styles(ws, 18)
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:R{ws.max_row}"
    _auto_width(ws, 1, 18)
    wb.save(full_path)

    # --- Sales WBS ---
    sales_df = df[sales_cols].rename(columns={"Remarks / Comments": "Remarks/Comments"})
    sales_df.to_excel(sales_path, sheet_name="WBS", index=False, startrow=4)
    swb = load_workbook(sales_path)
    sws = swb["WBS"]
    _write_header(sws, project_title, company_name, project_manager, today, 7)
    _insert_phase_headers(sws, 7)
    _dedup_column(sws, "A")
    _dedup_column(sws, "B")
    _apply_styles(sws, 7)
    sws.freeze_panes = "A6"
    sws.auto_filter.ref = f"A5:G{sws.max_row}"
    _auto_width(sws, 1, 7)
    swb.save(sales_path)

    return full_path, sales_path


def build_rows(tasks: list, phase_name: str, global_task_id: int, global_sub_task_id: int):
    rows = []
    for task in tasks:
        details = [d.strip() for d in task.get("detailed_information", "").split("|") if d.strip()] or [""]
        for i, detail in enumerate(details):
            first = i == 0
            rows.append({
                "Phase": phase_name if first else "",
                "Module Name": task.get("module_name", "") if first else "",
                "Task ID": global_task_id if first else "",
                "Task Title": task.get("task_title", "") if first else "",
                "Sub Task ID": global_sub_task_id,
                "Sub Task": detail,
                "Priority": task.get("priority", "") if first else "",
                "Dependency": task.get("dependency", "") if first else "",
                "Owner": task.get("owner", "") if first else "",
                "Status": task.get("status", "") if first else "",
                "Start Date": task.get("start_date", "") if first else "",
                "Working Duration (Days)": task.get("working_duration_days", "") if first else "",
                "Actual Completion Date": task.get("actual_completion_date", "") if first else "",
                "Delay (In Days)": task.get("delay_days", "") if first else "",
                "% Complete": task.get("percent_complete", "") if first else "",
                "Effort (Hours)": task.get("effort_hours", "") if first else "",
                "Sprint / Milestone": task.get("sprint_milestone", "") if first else "",
                "Remarks / Comments": task.get("remarks_comments", "") if first else "",
            })
            global_sub_task_id += 1
        rows.append({})  # spacer
        global_task_id += 1
    return rows, global_task_id, global_sub_task_id
