import asyncio
import json
import uuid
import os
from io import BytesIO
from datetime import datetime
from typing import List, Optional

import fitz
from fastapi import FastAPI, BackgroundTasks, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware

from models.schemas import (
    WBSRequest, JobResponse,
    ScopeRequest, ModuleRequest, PhaseRequest, TaskRequest,
)
from core.job_store import (
    create_job, get_job, get_queue,
    push_event, push_done, push_error, set_running,
)
from agents.scope_agent import run_scope_agent
from agents.module_agent import run_module_agent
from agents.phase_agent import run_phase_agent
from agents.task_agent import run_task_agent
from services.excel_builder import build_excel, build_rows
from services.mail_service import send_wbs_email
from db import queries
from core.config import settings

app = FastAPI(title="WBS AI System", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ─────────────────────────────────────────────
# BACKGROUND PIPELINE
# ─────────────────────────────────────────────

async def run_wbs_pipeline(job_id: str, req: WBSRequest):
    set_running(job_id)
    await queries.update_job_status(job_id, "running")

    project_config = {k: v.upper() for k, v in req.project_config.model_dump().items()}

    try:
        # STEP 1 – SCOPE
        await push_event(job_id, {"step": 1, "agent": "Scope Agent", "status": "running"})
        detailed_scope, _ = await run_scope_agent(req.rough_scope)
        await push_event(job_id, {"step": 1, "agent": "Scope Agent", "status": "done", "result": detailed_scope})

        # STEP 2 – MODULES
        await push_event(job_id, {"step": 2, "agent": "Module Agent", "status": "running"})
        module_json, _ = await run_module_agent(detailed_scope)
        module_names = [m["module_name"] for m in module_json["modules"]]
        await push_event(job_id, {"step": 2, "agent": "Module Agent", "status": "done", "result": module_names})

        # STEP 3 – PHASES
        await push_event(job_id, {"step": 3, "agent": "Phase Agent", "status": "running"})
        phase_json, _ = await run_phase_agent(detailed_scope, module_json, project_config)
        phases = phase_json.get("phases", [])
        phase_names = [p["phase_name"] for p in phases]
        await push_event(job_id, {"step": 3, "agent": "Phase Agent", "status": "done", "result": phase_names})

        # STEP 4 – TASKS
        all_rows = []
        global_task_id = 1
        global_sub_task_id = 1
        previous_context = ""
        total_phases = len(phases)

        for index, phase in enumerate(phases, start=1):
            phase_name = phase.get("phase_name", "")
            assigned_modules = phase.get("modules", [])

            await push_event(job_id, {"step": 4, "agent": "Task Agent", "status": "running", "phase": phase_name, "phase_index": index, "total_phases": total_phases})

            task_json, _ = await run_task_agent(
                detailed_scope=detailed_scope,
                modules=module_json,
                project_config=project_config,
                phase_name=phase_name,
                assigned_modules=assigned_modules,
                previous_context=previous_context,
                team_size=req.team_size,
                project_start_date=req.project_start_date,
            )

            tasks = task_json.get("tasks", [])
            rows, global_task_id, global_sub_task_id = build_rows(
                tasks, phase_name, global_task_id, global_sub_task_id
            )
            all_rows.extend(rows)

            previous_context += f"\nPHASE: {phase_name}\n"
            for t in tasks:
                previous_context += f"  Module: {t.get('module_name')} | Task: {t.get('task_title')}\n"

            await push_event(job_id, {"step": 4, "agent": "Task Agent", "status": "done", "phase": phase_name, "phase_index": index, "total_phases": total_phases, "task_count": len(tasks)})

        # BUILD EXCEL
        await push_event(job_id, {"step": 5, "agent": "Excel Builder", "status": "running"})
        full_path, sales_path = build_excel(
            all_rows,
            req.project_title,
            req.company_name,
            req.project_manager,
            job_id,
        )
        await push_event(job_id, {"step": 5, "agent": "Excel Builder", "status": "done"})

        # EMAIL
        if req.recipient_emails:
            await push_event(job_id, {"step": 6, "agent": "Email Service", "status": "running", "recipient": ", ".join(req.recipient_emails)})
            await asyncio.to_thread(
                send_wbs_email,
                [str(e) for e in req.recipient_emails],
                [str(e) for e in req.cc_emails],
                req.project_title,
                full_path,
                sales_path,
            )
            await push_event(job_id, {"step": 6, "agent": "Email Service", "status": "done", "recipient": ", ".join(req.recipient_emails)})

        await queries.update_job_status(job_id, "done")
        await push_done(job_id, {"full_wbs_local": full_path, "sales_wbs_local": sales_path})

    except Exception as e:
        error_msg = str(e)
        await queries.update_job_status(job_id, "failed", error=error_msg)
        await push_error(job_id, error_msg)


# ─────────────────────────────────────────────
# ROUTES
# ─────────────────────────────────────────────

async def submit_wbs_job(req: WBSRequest, background_tasks: BackgroundTasks) -> JobResponse:
    job_id = str(uuid.uuid4())
    create_job(job_id)
    await queries.create_job(job_id, {
        "project_title": req.project_title,
        "company_name": req.company_name,
        "project_manager": req.project_manager,
        "team_size": req.team_size,
        "project_start_date": req.project_start_date,
        "rough_scope": req.rough_scope,
        "project_config": req.project_config.model_dump(),
        "recipient_emails": [str(e) for e in req.recipient_emails],
        "cc_emails": [str(e) for e in req.cc_emails],
    })
    background_tasks.add_task(run_wbs_pipeline, job_id, req)
    return JobResponse(job_id=job_id, message="WBS generation started. Connect to /status/{job_id} for live updates.")


async def extract_supporting_documents_text(files: Optional[List[UploadFile]]) -> str:
    if not files:
        return ""

    chunks = []
    for file in files:
        filename = file.filename or "document"
        content = await file.read()
        if not content:
            continue

        lower_name = filename.lower()
        if lower_name.endswith(".pdf") or file.content_type == "application/pdf":
            try:
                doc = fitz.open(stream=content, filetype="pdf")
                text = "\n".join(page.get_text().strip() for page in doc)
                doc.close()
            except Exception as exc:
                raise HTTPException(status_code=400, detail=f"Could not read PDF: {filename}") from exc
        elif lower_name.endswith(".txt") or (file.content_type or "").startswith("text/"):
            text = content.decode("utf-8", errors="ignore")
        elif lower_name.endswith(".xlsx"):
            try:
                text = extract_excel_text(content)
            except Exception as exc:
                raise HTTPException(status_code=400, detail=f"Could not read Excel file: {filename}") from exc
        else:
            continue

        if text.strip():
            chunks.append(f"SUPPORTING DOCUMENT - {filename}:\n{text.strip()}")

    return "\n\n".join(chunks)


def extract_excel_text(content: bytes, max_rows_per_sheet: int = 500) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    chunks = []

    for sheet in workbook.worksheets:
        rows = []
        for row in sheet.iter_rows(values_only=True):
            values = [format_cell_value(value) for value in row]
            while values and not values[-1]:
                values.pop()
            if any(values):
                rows.append(values)

        if not rows:
            continue

        headers = rows[0]
        data_rows = rows[1:]
        included_rows = data_rows[:max_rows_per_sheet]
        max_cols = max(len(headers), *(len(row) for row in included_rows)) if included_rows else len(headers)
        headers = pad_row(headers, max_cols)

        lines = [
            f"Sheet: {sheet.title}",
            f"Columns: {' | '.join(headers)}",
            "Rows:",
        ]

        for index, row in enumerate(included_rows, start=1):
            lines.append(f"{index}. {' | '.join(pad_row(row, max_cols))}")

        if len(data_rows) > max_rows_per_sheet:
            lines.append(f"Only first {max_rows_per_sheet} rows included from {len(data_rows)} total rows.")

        chunks.append("\n".join(lines))

    workbook.close()
    return "\n\n".join(chunks)


def format_cell_value(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def pad_row(row: List[str], length: int) -> List[str]:
    return row + [""] * (length - len(row))


@app.post("/generate-wbs", response_model=JobResponse, tags=["WBS Pipeline"])
async def generate_wbs(req: WBSRequest, background_tasks: BackgroundTasks):
    """Submit a WBS generation job. Returns job_id immediately."""
    return await submit_wbs_job(req, background_tasks)


@app.post("/generate-wbs-with-documents", response_model=JobResponse, tags=["WBS Pipeline"])
async def generate_wbs_with_documents(
    background_tasks: BackgroundTasks,
    project_title: str = Form(...),
    company_name: str = Form(...),
    project_manager: str = Form(...),
    team_size: int = Form(...),
    project_start_date: str = Form(...),
    rough_scope: str = Form(...),
    project_config: str = Form(...),
    recipient_emails: str = Form(...),
    cc_emails: str = Form("[]"),
    supporting_documents: Optional[List[UploadFile]] = File(None),
):
    """Submit a WBS job with optional supporting PDF/TXT documents."""
    try:
        project_config_data = json.loads(project_config)
        recipient_email_data = json.loads(recipient_emails)
        cc_email_data = json.loads(cc_emails)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="Invalid JSON form payload") from exc

    document_text = await extract_supporting_documents_text(supporting_documents)
    combined_scope = rough_scope.strip()
    if document_text:
        combined_scope = f"{combined_scope}\n\nSUPPORTING DOCUMENTS:\n{document_text}"

    req = WBSRequest(
        project_title=project_title,
        company_name=company_name,
        project_manager=project_manager,
        team_size=team_size,
        project_start_date=project_start_date,
        rough_scope=combined_scope,
        project_config=project_config_data,
        recipient_emails=recipient_email_data,
        cc_emails=cc_email_data,
    )
    return await submit_wbs_job(req, background_tasks)


@app.get("/status/{job_id}", tags=["WBS Pipeline"])
async def stream_status(job_id: str):
    """SSE stream for live progress updates of a job."""
    job = get_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    queue = get_queue(job_id)

    async def event_generator():
        for payload in job.get("events", []):
            yield f"event: progress\ndata: {json.dumps(payload)}\n\n"

        if job["status"] in ("done", "failed"):
            if job["status"] == "done":
                yield f"event: done\ndata: {json.dumps(job['result'])}\n\n"
            else:
                yield f"event: error\ndata: {json.dumps({'message': job['error']})}\n\n"
            return

        while True:
            try:
                event = await asyncio.wait_for(queue.get(), timeout=60)
                yield f"event: {event['event']}\ndata: {json.dumps(event['data'])}\n\n"
                if event["event"] in ("done", "error"):
                    break
            except asyncio.TimeoutError:
                yield f"event: ping\ndata: {{}}\n\n"

    return StreamingResponse(event_generator(), media_type="text/event-stream")


@app.get("/job/{job_id}", tags=["WBS Pipeline"])
async def get_job_details(job_id: str):
    """Get job input + status from MongoDB."""
    job = await queries.get_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    job["_id"] = str(job["_id"])
    return job


@app.get("/download/{job_id}/{file_type}", tags=["WBS Pipeline"])
async def download_file(job_id: str, file_type: str):
    """Direct download: file_type = full_wbs or sales_wbs"""
    job = get_job(job_id)
    if not job or not job.get("result"):
        raise HTTPException(status_code=404, detail="Job not ready or not found")

    key = "full_wbs_local" if file_type == "full_wbs" else "sales_wbs_local"
    path = job["result"].get(key)

    if not path or not os.path.exists(path):
        raise HTTPException(status_code=404, detail="File not found on server")

    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=os.path.basename(path),
    )


# ─────────────────────────────────────────────
# INDIVIDUAL AGENT ENDPOINTS
# ─────────────────────────────────────────────

@app.post("/agents/scope", tags=["Individual Agents"])
async def agent_scope(req: ScopeRequest):
    content, tokens = await run_scope_agent(req.rough_scope)
    return {"scope": content, "tokens": tokens}


@app.post("/agents/modules", tags=["Individual Agents"])
async def agent_modules(req: ModuleRequest):
    modules, tokens = await run_module_agent(req.detailed_scope)
    return {"modules": modules, "tokens": tokens}


@app.post("/agents/phases", tags=["Individual Agents"])
async def agent_phases(req: PhaseRequest):
    phases, tokens = await run_phase_agent(
        req.detailed_scope,
        req.modules,
        req.project_config.model_dump(),
    )
    return {"phases": phases, "tokens": tokens}


@app.post("/agents/tasks", tags=["Individual Agents"])
async def agent_tasks(req: TaskRequest):
    phases = req.phases.get("phases", [])
    all_tasks = []
    previous_context = ""
    for phase in phases:
        task_json, tokens = await run_task_agent(
            detailed_scope=req.detailed_scope,
            modules=req.modules,
            project_config=req.project_config.model_dump(),
            phase_name=phase["phase_name"],
            assigned_modules=phase["modules"],
            previous_context=previous_context,
            team_size=req.team_size,
            project_start_date=req.project_start_date,
        )
        all_tasks.append({"phase": phase["phase_name"], "tasks": task_json, "tokens": tokens})
        for t in task_json.get("tasks", []):
            previous_context += f"  Module: {t.get('module_name')} | Task: {t.get('task_title')}\n"
    return {"phases_tasks": all_tasks}


@app.get("/managers", tags=["Settings"])
async def get_managers():
    return await queries.get_managers()

@app.post("/managers", tags=["Settings"])
async def add_manager(payload: dict):
    name = payload.get("name", "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Name is required")
    await queries.add_manager(name)
    return {"message": "Manager added"}

@app.delete("/managers/{name}", tags=["Settings"])
async def delete_manager(name: str):
    await queries.delete_manager(name)
    return {"message": "Manager deleted"}


@app.get("/emails", tags=["Settings"])
async def get_emails():
    return await queries.get_emails()

@app.post("/emails", tags=["Settings"])
async def add_email(payload: dict):
    email = payload.get("email", "").strip()
    if not email:
        raise HTTPException(status_code=400, detail="Email is required")
    await queries.add_email(email)
    return {"message": "Email added"}

@app.delete("/emails/{email}", tags=["Settings"])
async def delete_email(email: str):
    await queries.delete_email(email)
    return {"message": "Email deleted"}


@app.get("/jobs", tags=["WBS Pipeline"])
async def list_jobs(page: int = 1, limit: int = 10):
    """Get paginated jobs from MongoDB."""
    skip = (page - 1) * limit
    return await queries.list_jobs(skip=skip, limit=limit)


@app.get("/config", tags=["System"])
async def get_config():
    return {"api_base_url": settings.API_BASE_URL}


@app.get("/health", tags=["System"])
async def health():
    return {"status": "ok", "timestamp": datetime.utcnow().isoformat()}
